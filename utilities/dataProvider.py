import os

import openpyxl

def get_data(sheetName, fileName="testdata.xlsx"):
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Directory where configReader.py is located
    EXCEL_PATH = os.path.join(BASE_DIR, "..", "exceldata", fileName)

    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows + 1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.append(data)
        if any(v is not None for v in dataList):
            mainList.append(dataList)
    return mainList

