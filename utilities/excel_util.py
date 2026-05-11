import os

import openpyxl

def get_row_count(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def get_col_count(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def get_cell_data(path,sheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum,column=colNum).value


def set_cell_data(path,sheetName,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum,column=colNum).value=data
    workbook.save(path)


def get_col_by_header(path, sheetName, headerName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == headerName:
            return col
    raise ValueError(f"Header '{headerName}' not found in sheet '{sheetName}'")